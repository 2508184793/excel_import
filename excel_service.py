from openpyxl import load_workbook  # 导入openpyxl库中的load_workbook函数，用于加载Excel文件

"""
获取Excel文件的基本信息，包括所有工作表名称和当前活动工作表名称。

参数:
    file_path (str): Excel文件的路径

返回:
    dict: 包含sheet_names（所有工作表名称列表）和active_sheet（当前活动工作表名称）的字典
"""


def get_excel_info(file_path):
    wb = load_workbook(file_path)  # 加载Excel工作簿
    sheet_names = wb.sheetnames  # 获取所有工作表的名称，返回一个列表
    active_sheet = wb.active.title  # 获取当前活动工作表的名称

    # 打印工作簿的基本信息
    print(f"工作簿包含的工作表: {sheet_names}")
    print(f"当前活动工作表: {active_sheet}")

    return {
        "sheet_names": sheet_names,  # 所有工作表名称
        "active_sheet": active_sheet  # 当前活动工作表名称
    }


from typing import Dict, Any  # 导入类型注解

"""
获取指定工作表中指定行的所有单元格内容（通常用于获取标题行）。
参数:
    sheet_name (str): 工作表名称
    row_index (int): 行索引（从1开始）
    file_path (str): Excel文件路径
返回:
    dict: 包含'title_row'键，对应该行所有单元格的值列表
"""
def get_title_info(sheet_name: str, row_index: int, file_path):
    wb = load_workbook(file_path)  # 加载Excel工作簿
    ws = wb[sheet_name]  # 获取指定名称的工作表
    # 获取指定行的所有单元格的值，返回一个列表
    return {
        "title_row": [cell.value for cell in ws[row_index]]
    }

"""
加载Excel文件并打印当前活动工作表的所有内容。
参数:
    file_name (str): Excel文件名（不包含路径）
返回:
    None
"""
def load_excel(file_name):
    import os
    file_path = os.path.join(os.getcwd(), "FileDir", file_name)  # 构建完整文件路径
    wb = load_workbook(file_path)  # 加载Excel工作簿
    ws = wb.active  # 获取当前活动工作表
    # 遍历工作表的每一行
    for row in ws.rows:
        # 将该行每个单元格的值转换为字符串，组成一个列表
        row_values = [str(cell.value) for cell in row]
        # 打印该行的所有单元格内容，用空格分隔
        print(" ".join(row_values))
